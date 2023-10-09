from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.colors import SchemeColor
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame
from copy import deepcopy

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows

from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import GradientFillProperties, GradientStop

from config import settings
from data_preparation.chart_date_title import DateTitle


async def create_fig(df_action: DataFrame, df_categories: DataFrame,
                     sheet_name: str = settings.WEEKLY_XLSX_FILE_NAME) -> None:
    """
    The create_fig function creates a bar chart and pie chart from the dataframes
        passed to it.

    :param df_action: DataFrame: Pass the dataframe to the function
    :param df_categories: DataFrame: Pass the dataframe to the function
    :param sheet_name: str: Specify the name of the file that will be created
    :return: Nothing, but it creates a workbook object
    """
    wb: Workbook = Workbook()
    create_bar_charts(df=df_action, wb=wb)
    create_pie_chart(df=df_categories, wb=wb)
    wb.save(sheet_name)


def create_bar_charts(*, df: DataFrame, wb: Workbook) -> None:
    """
    The create_bar_charts function creates a bar chart for each action in the dataframe.

    :param df: DataFrame: Pass the dataframe to the function
    :param wb: Workbook: Pass the workbook object to the function
    :return: None
    """
    ws: Worksheet = wb.active
    ws.title = f"bars(week - {DateTitle.week})"
    for row in dataframe_to_rows(df, index=True, header=True):
        ws.append(row)
    ws.delete_rows(2)
    ws['A1'] = "dow/actions"
    ws.column_dimensions["A"].width = 14
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrapText=True)
    chart = BarChart()
    chart.y_axis.title = 'Devoted time'
    chart.x_axis.title = 'Days of the week'

    data = Reference(ws, min_col=2, min_row=1, max_row=8, max_col=len(df.columns) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=8)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 6
    simple_stack_chart(chart, ws)
    horizontal_bar_chart(chart, ws)
    stacked_chart(chart, ws)
    percent_stacked_chart(chart, ws)


def simple_stack_chart(chart: BarChart, ws: Worksheet) -> None:
    """
    Add to wb simple bar chart

    :param chart: BarChart: Pass the chart object to the function
    :param ws: Worksheet: Specify the worksheet that the chart will be added to
    :return: A simple bar chart
    """
    chart1 = deepcopy(chart)
    chart1.type = "col"
    chart1.style = 10
    chart1.title = DateTitle.get_current_dow()
    ws.add_chart(chart1, "A10")


def horizontal_bar_chart(chart: BarChart, ws: Worksheet) -> None:
    """
    Add to wb horizontal bar chart

    :param chart: BarChart: Pass the chart object to the function
    :param ws: Worksheet: Specify the worksheet that the chart will be added to
    :return: A horizontal bar chart
    """
    chart2 = deepcopy(chart)
    chart2.style = 10
    chart2.type = "bar"
    chart2.title = DateTitle.get_current_dow()
    ws.add_chart(chart2, "K10")


#
def stacked_chart(chart: BarChart, ws: Worksheet) -> None:
    """
    Add to wb stacked chart

    :param chart: BarChart: Pass the chart object to the function
    :param ws: Worksheet: Specify the worksheet that the chart will be added to
    :return: A stacked bar chart
    """
    chart3 = deepcopy(chart)
    chart3.type = "col"
    chart3.style = 10
    chart3.grouping = "stacked"
    chart3.overlap = 100
    chart3.title = DateTitle.get_current_dow()
    ws.add_chart(chart3, "A27")


def percent_stacked_chart(chart: BarChart, ws: Worksheet) -> None:
    """
    Add to wb percent stacked chart

    :param chart: BarChart: Pass the chart object to the function
    :param ws: Worksheet: Specify the worksheet that the chart will be added to
    :return: A percent stacked bar chart
    """
    chart4 = deepcopy(chart)
    chart4.type = "bar"
    chart4.style = 10
    chart4.grouping = "percentStacked"
    chart4.overlap = 100
    chart4.title = DateTitle.get_current_dow()
    ws.add_chart(chart4, "K27")


def create_pie_chart(*, df: DataFrame, wb: Workbook) -> None:
    """
    The create_pie_chart function creates a pie chart of the dataframe passed to it.

    :param df: DataFrame: Pass a dataframe to the function
    :param wb: Workbook: Create a new worksheet in the workbook
    :return: None
    """
    ws: Worksheet = wb.create_sheet(title=f"pie(week - {DateTitle.week})")
    for row in dataframe_to_rows(df, index=True, header=True):
        ws.append(row)
    ws.delete_rows(2)
    ws['A1'] = "quantity"
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10
    pie = PieChart()
    labels = Reference(ws, min_col=2, min_row=2, max_row=len(df) + 1, max_col=3)
    data = Reference(ws, min_col=1, min_row=1, max_row=len(df) + 1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Amount of time spent by category"
    gradient_pie(pie, ws)


def gradient_pie(pie: PieChart, ws: Worksheet) -> None:
    """
    Add to wb gradient pie chart

    :param pie: PieChart: Pass in the pie chart object
    :param ws: Worksheet: Add the chart to a specific worksheet
    :return: None
    """
    # Cut the first slice out of the pie and apply a gradient to it
    slice_ = DataPoint(idx=0, explosion=20, spPr=GraphicalProperties(
        gradFill=GradientFillProperties(
            gsLst=(
                GradientStop(
                    pos=0,
                    prstClr='blue'
                ),
                GradientStop(
                    pos=100000,
                    schemeClr=SchemeColor(
                        val='accent1',
                        lumMod=30000,
                        lumOff=70000
                    )
                )
            )
        )
    )
                      )
    pie.series[0].data_points = [slice_]
    ws.add_chart(pie, "E1")
