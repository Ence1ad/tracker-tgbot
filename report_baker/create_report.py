from openpyxl.chart.marker import DataPoint
from openpyxl.drawing.colors import SchemeColor
from openpyxl.styles import Alignment
from openpyxl.worksheet.worksheet import Worksheet
from pandas import DataFrame
from copy import deepcopy

from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference, PieChart
from openpyxl.utils.dataframe import dataframe_to_rows

from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.drawing.fill import GradientFillProperties, GradientStop

from config import settings
from report_baker.chart_date_title import DateTitle


async def create_fig(df_action: DataFrame, df_categories: DataFrame,
                     file_name: str = settings.WEEKLY_XLSX_FILE_NAME) -> None:
    """Create figures and charts in an Excel workbook.

    :param df_action: DataFrame: DataFrame containing action data.
    :param df_categories: DataFrame: DataFrame containing category data.
    :param file_name: str: Name of the Excel file to save the figures and charts
    (default: settings.WEEKLY_XLSX_FILE_NAME).
    :return: None
    """
    wb: Workbook = Workbook()
    create_bar_charts(df=df_action, wb=wb)
    create_pie_chart(df=df_categories, wb=wb)
    wb.save(file_name)


def create_bar_charts(*, df: DataFrame, wb: Workbook) -> None:
    """Create bar charts in the Excel workbook.

    :param df: DataFrame: DataFrame containing data for the bar charts.
    :param wb: Workbook: Excel workbook to add the charts.
    :return: None
    """
    ws: Worksheet = wb.active
    ws.title = f"bars(week - {DateTitle.week})"
    for row in dataframe_to_rows(df, index=True, header=True):
        ws.append(row)
    ws.delete_rows(2)
    ws['A1'] = "dow/actions"
    ws.column_dimensions["A"].width = 14
    for row in ws.iter_rows():
        for cell in row:
            cell.alignment = Alignment(wrapText=True)
    chart = BarChart()
    chart.y_axis.title = 'Devoted time'
    chart.x_axis.title = 'Days of the week'

    data = Reference(ws, min_col=2, min_row=1, max_row=8, max_col=len(df.columns) + 1)
    cats = Reference(ws, min_col=1, min_row=2, max_row=8)
    chart.add_data(data, titles_from_data=True)
    chart.set_categories(cats)
    chart.shape = 6
    simple_stack_chart(chart, ws)
    horizontal_bar_chart(chart, ws)
    stacked_chart(chart, ws)
    percent_stacked_chart(chart, ws)


def simple_stack_chart(chart: BarChart, ws: Worksheet) -> None:
    """Create a horizontal bar chart and add it to a worksheet.

    :param chart: BarChart: The bar chart to modify.
    :param ws: Worksheet: The worksheet to add the chart to.
    :return: None
    """
    chart1 = deepcopy(chart)
    chart1.type = "col"
    chart1.style = 10
    chart1.title = DateTitle.get_current_dow()
    ws.add_chart(chart1, "A10")


def horizontal_bar_chart(chart: BarChart, ws: Worksheet) -> None:
    """Create a horizontal bar chart and add it to a worksheet.

    :param chart: BarChart: The bar chart to modify.
    :param ws: Worksheet: The worksheet to add the chart to.
    :return: None
    """
    chart2 = deepcopy(chart)
    chart2.style = 10
    chart2.type = "bar"
    chart2.title = DateTitle.get_current_dow()
    ws.add_chart(chart2, "K10")


def stacked_chart(chart: BarChart, ws: Worksheet) -> None:
    """Create a bar chart with a stacked layout and add it to a worksheet.

    :param chart: BarChart: The bar chart to modify.
    :param ws: Worksheet: The worksheet to add the chart to.
    :return: None
    """
    chart3 = deepcopy(chart)
    chart3.type = "col"
    chart3.style = 10
    chart3.grouping = "stacked"
    chart3.overlap = 100
    chart3.title = DateTitle.get_current_dow()
    ws.add_chart(chart3, "A27")


def percent_stacked_chart(chart: BarChart, ws: Worksheet) -> None:
    """Create a bar chart with a percent stacked layout and add it to a worksheet.

    :param chart: BarChart: The bar chart to modify.
    :param ws: Worksheet: The worksheet to add the chart to.
    :return: None
    """
    chart4 = deepcopy(chart)
    chart4.type = "bar"
    chart4.style = 10
    chart4.grouping = "percentStacked"
    chart4.overlap = 100
    chart4.title = DateTitle.get_current_dow()
    ws.add_chart(chart4, "K27")


def create_pie_chart(*, df: DataFrame, wb: Workbook) -> None:
    """Create a pie chart in the Excel workbook.

    :param df: DataFrame: DataFrame containing data for the pie chart.
    :param wb: Workbook: Excel workbook to add the chart.
    :return:  None
    """
    ws: Worksheet = wb.create_sheet(title=f"pie(week - {DateTitle.week})")
    for row in dataframe_to_rows(df, index=True, header=True):
        ws.append(row)
    ws.delete_rows(2)
    ws['A1'] = "quantity"
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10
    pie = PieChart()
    labels = Reference(ws, min_col=2, min_row=2, max_row=len(df) + 1, max_col=3)
    data = Reference(ws, min_col=1, min_row=1, max_row=len(df) + 1)
    pie.add_data(data, titles_from_data=True)
    pie.set_categories(labels)
    pie.title = "Amount of time spent by category"
    gradient_pie(pie, ws)


def gradient_pie(pie: PieChart, ws: Worksheet) -> None:
    """Apply a gradient to the first slice of a pie chart and add it to a worksheet.

    :param pie: PieChart: The pie chart to modify.
    :param ws: Worksheet: The worksheet to add the chart to.
    :return: None
    """
    # Cut the first slice out of the pie and apply a gradient to it
    slice_ = DataPoint(idx=0, explosion=20, spPr=GraphicalProperties(
        gradFill=GradientFillProperties(
            gsLst=(
                GradientStop(
                    pos=0,
                    prstClr='blue'
                ),
                GradientStop(
                    pos=100000,
                    schemeClr=SchemeColor(
                        val='accent1',
                        lumMod=30000,
                        lumOff=70000
                    )
                )
            )
        )
    )
                      )
    pie.series[0].data_points = [slice_]
    ws.add_chart(pie, "E1")
